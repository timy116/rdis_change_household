import datetime
import json
import openpyxl
import os
import time
from collections import namedtuple
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side

MAIN = True
# SAMPLE_PATH = '..\\..\\input\\easy.txt'
SAMPLE_PATH = '..\\..\\input\\main_107farmerSurvey_investigator.txt' if MAIN else '..\\..\\input\\sub_107farmerSurvey_investigator.txt'
JSON_PATH = '..\\..\\output\\json\\公務資料.json' if MAIN else '..\\..\\output\\json\\公務資料_備選.json'
# JSON_PATH = '..\\..\\output\\json\\json.json'
FOLDER_NAME = '主選特約_公務資料' if MAIN else '備選特約_公務資料'
FOLDER_PATH = '..\\..\\output\\' + datetime.datetime.now().strftime('%Y%m%d_%H%M%S') + FOLDER_NAME

EXCEPT_NUM = ['100140500520', '640200004190', '100091202591', '100140407561', '670090000233', '100091700643',
              '660290006014', '670320007154',
              '670260004055', '100131304635']
SAMPLE_TITLES = ['農戶編號', '調查姓名', '電話', '地址', '出生年', '原層別', '連結編號']
HOUSEHOLD_TITLES = ['[戶籍檔]', '出生年', '關係', '死亡或除戶', '農保', '老農津貼', '國保給付', '勞保給付', '勞退給付', '農保給付']
TRANSFER_CROP_TITLES = ['[轉作補貼]', '項目', '作物名稱', '金額', '期別']
DISASTER_TITLES = ['[災害]', '項目', '災害', '核定作物', '核定面積', '金額']
SB_SBDY_TITLES = ['[107小大]', '姓名', '大專業農轉契作', '小地主出租給付', '離農獎勵']
LIVESTOCK_TITLES = ['[畜牧資訊]', '年', '調查時間', '畜牧品項', '在養數量', '屠宰數量', '副產品名稱', '副產品數量']
SAMPLE_ROSTER_TITLES = ['序號', '樣本套號 ', '農戶編號', '連結編號 ', '戶長姓名', '電話 ', '地址 ', '層別 ', '經營種類 ', '可耕作地面積', '成功打勾']
SAMPLE_ATTR = [
    'layer',
    'name',
    'tel',
    'addr',
    'county',
    'town',
    'link_num',
    'id',
    'num',
    'main_type',
    'area',
    'sample_num',
    'inv_name'
]
Sample = namedtuple('Sample', SAMPLE_ATTR)

TYPE_FLAG = '主選' if MAIN else '備選'
ALIGNMENT = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
SIDE = Side(style='medium')
BORDER = Border(
    top=SIDE,
    bottom=SIDE,
    left=SIDE,
    right=SIDE
)

# sorted by county
sample_dict = {}
investigator_dict = {}
official_data = json.loads(open(JSON_PATH, encoding='utf8').read())

if not os.path.isdir(FOLDER_PATH):
    os.mkdir(FOLDER_PATH)


def set_excel_title(sheet, row_index, flag, titles) -> None:
    if flag == 'sample_roster':
        for index, title in enumerate(titles, start=1):
            cell = sheet.cell(row_index, index)
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            cell.value = title
            cell.border = BORDER
    else:
        for index, title in enumerate(titles, start=1):
            sheet.cell(column=index, row=row_index).value = title


def read_sample() -> None:
    """
    讀取 sample 檔並使用 dict, key = county : value = 住在這縣市的人
    """
    with open(SAMPLE_PATH, encoding='utf8') as f:
        for line in f:
            sample = Sample._make(line.split('\t'))
            inv_name = sample.inv_name.strip()

            if inv_name not in investigator_dict:
                investigate_l = [sample]
                investigator_dict[inv_name] = investigate_l
            else:
                investigator_dict.get(inv_name).append(sample)


def output_excel(type_flag=TYPE_FLAG) -> None:
    for inv_name, samples in investigator_dict.items():
        if type_flag == '主選':
            samples.sort(key=lambda x: x.county + x.town + x.addr)
        else:
            samples.sort(key=lambda x: x.num[-5:])
        wb = openpyxl.Workbook()
        col_index = 1
        row_index = 1
        sheet = wb.active
        sheet.title = inv_name

        for sample in samples:
            scholarship = ''
            sb = ''
            farmer_num = sample.num
            crops = []
            sample_data = official_data.get(farmer_num)
            if sample_data is None:
                if farmer_num in EXCEPT_NUM:
                    ...
                else:
                    ...
            else:
                if row_index - 1 == 0:
                    width = list(
                        map(lambda x: x * 1.054, [14.29, 9.29, 16.29, 29.29, 9.29, 11.29, 11.29, 11.29, 11.29]))
                    for i in range(1, len(width) + 1):
                        sheet.column_dimensions[get_column_letter(i)].width = width[i - 1]

                set_excel_title(sheet, row_index, 'sample', SAMPLE_TITLES)
                row_index += 1
                info = [
                    farmer_num, sample_data.get('name'), sample_data.get('telephone'), sample_data.get('address'),
                    sample_data.get('birthday'), sample_data.get('layer'), sample_data.get('serial')
                ]
                for index, value in enumerate(info, start=1):
                    sheet.cell(column=index, row=row_index).value = value
                    sheet.cell(column=index, row=row_index).alignment = Alignment(wrap_text=True)

                row_index += 1
                separate_start = ' ' + '-'*206 + ' '
                sheet.cell(column=col_index, row=row_index).value = separate_start

                row_index += 1
                set_excel_title(sheet, row_index, 'household', HOUSEHOLD_TITLES)
                household = sample_data.get('household')
                household.sort(key=lambda x: x[1])

                for person in household:
                    row_index += 1
                    for index, p_data in enumerate(person, start=2):
                        if index in [5 + 2, 6 + 2, 7 + 2, 8 + 2] and p_data:
                            sheet.cell(column=index, row=row_index).number_format = '#,###,###'
                            p_data = eval(p_data.replace(',', ''))
                        if index == 9 + 2:
                            if person[9]:
                                scholarship += person[9] + ','
                            continue
                        if index == 10 + 2 and person[10] not in sb:
                            sb += person[10]
                            break
                        sheet.cell(column=index, row=row_index).value = p_data
                        sheet.cell(column=index, row=row_index).alignment = Alignment(horizontal='left')

                # 輸出申報核定資料，檢查是否有資料
                declaration = sample_data.get('declaration')
                if declaration:
                    row_index += 2
                    sheet.cell(column=1, row=row_index).value = '[申報核定]'
                    sheet.cell(column=2, row=row_index).value = declaration

                # 輸出轉作補貼資料，檢查是否有資料
                crop_sbdy = sample_data.get('cropSbdy')
                if crop_sbdy:
                    row_index += 2
                    crop_d = {}
                    for i in crop_sbdy:
                        crop_name = i[0]
                        amount = int(eval(i[1]))
                        if crop_name not in crop_d:
                            crop_d[crop_name] = amount
                        else:
                            crop_d[crop_name] = crop_d.get(crop_name) + amount

                    item_index = 0
                    set_excel_title(sheet, row_index, 'transfer_crop', TRANSFER_CROP_TITLES)

                    for k, v in crop_d.items():
                        row_index += 1
                        item_index += 1
                        sheet.cell(column=2, row=row_index).value = item_index
                        sheet.cell(column=2, row=row_index).alignment = Alignment(horizontal='left')
                        sheet.cell(column=3, row=row_index).value = k

                        if len(k) > 8:
                            sheet.cell(column=3, row=row_index).alignment = Alignment(wrap_text=True)

                        if v:
                            sheet.cell(column=4, row=row_index).number_format = '#,###,###'
                        sheet.cell(column=4, row=row_index).value = v
                        sheet.cell(column=4, row=row_index).alignment = Alignment(horizontal='left')
                        sheet.cell(column=5, row=row_index).value = '1'

                        if k not in crops:
                            crops.append(k)

                # 輸出災害補助資料，檢查是否有資料
                disaster = sample_data.get('disaster')
                if disaster:
                    row_index += 1
                    item_index = 0
                    disaster_d = {}
                    for i in disaster:
                        data = {}
                        disaster_name = i[0] + '-' + i[1]
                        area = float(i[2])
                        amount = int(i[3])
                        if disaster_name not in disaster_d:
                            data['area'] = area
                            data['amount'] = amount
                        else:
                            data = disaster_d.get(disaster_name)
                            data['area'] = data.get('area') + area
                            data['amount'] = data.get('amount') + amount
                        disaster_d[disaster_name] = data

                    row_index += 1
                    set_excel_title(sheet, row_index, 'disaster', DISASTER_TITLES)

                    for k, v in disaster_d.items():
                        row_index += 1
                        item_index += 1
                        sheet.cell(column=2, row=row_index).value = item_index
                        sheet.cell(column=2, row=row_index).alignment = Alignment(horizontal='left')
                        l = k.split('-')
                        sheet.cell(column=3, row=row_index).value = l[0]

                        if len(l[0]) > 8:
                            sheet.cell(column=3, row=row_index).alignment = Alignment(wrap_text=True)
                        sheet.cell(column=4, row=row_index).value = l[1]
                        sheet.cell(column=5, row=row_index).value = v.get('area')
                        sheet.cell(column=5, row=row_index).alignment = Alignment(horizontal='left')

                        if v.get('amount'):
                            sheet.cell(column=6, row=row_index).number_format = '#,###,###'
                        sheet.cell(column=6, row=row_index).value = v.get('amount')
                        sheet.cell(column=6, row=row_index).alignment = Alignment(horizontal='left')

                        if l[1] not in crops:
                            crops.append(l[1])

                # 輸出小大補助資料，檢查是否有資料
                sb_sbdy = sample_data.get('sbSbdy')
                if sb_sbdy:
                    row_index += 2
                    set_excel_title(sheet, row_index, 'sb_sbdy', SB_SBDY_TITLES)
                    for i in sb_sbdy:
                        row_index += 1
                        for index, j in enumerate(i, start=2):
                            if index >= 3:
                                if int(eval(j)):
                                    sheet.cell(column=index, row=row_index).number_format = '#,###,###'
                                sheet.cell(column=index, row=row_index).alignment = Alignment(horizontal='left')
                                sheet.cell(column=index, row=row_index).value = int(eval(j))
                            else:
                                sheet.cell(column=index, row=row_index).value = j

                # 輸出畜牧資料，檢查是否有資料
                livestock = sample_data.get('livestock')
                if livestock:
                    row_index += 2
                    set_excel_title(sheet, row_index, 'livestock', LIVESTOCK_TITLES)
                    for k, v in livestock.items():
                        row_index += 1
                        sheet.cell(column=1, row=row_index).value = k
                        if len(k) > 6:
                            sheet.cell(column=1, row=row_index).alignment = Alignment(wrap_text=True)
                        v.sort(key=lambda x: x[6] + x[0])

                        for index, i in enumerate(v, start=1):
                            sheet.cell(column=2, row=row_index).value = i[6]
                            sheet.cell(column=3, row=row_index).value = i[0]
                            sheet.cell(column=4, row=row_index).value = i[1]
                            raise_count = int(eval(i[2])) if i[2].isnumeric() else i[2]
                            sheet.cell(column=5, row=row_index).value = raise_count
                            sheet.cell(column=5, row=row_index).alignment = Alignment(horizontal='left')
                            slaughter_count = int(eval(i[3])) if i[3].isnumeric() else i[3]
                            sheet.cell(column=6, row=row_index).value = slaughter_count
                            sheet.cell(column=6, row=row_index).alignment = Alignment(horizontal='left')
                            sheet.cell(column=7, row=row_index).value = i[4]
                            sheet.cell(column=7, row=row_index).alignment = Alignment(horizontal='left')
                            sheet.cell(column=8, row=row_index).value = int(eval(i[5]))
                            sheet.cell(column=8, row=row_index).alignment = Alignment(horizontal='left')
                            if index != len(v):
                                row_index += 1

                # 年度作物
                if crops:
                    row_index += 2
                    sheet.cell(column=1, row=row_index).value = '[106y-107y作物]'
                    sheet.cell(column=2, row=row_index).value = ','.join(crops)

                # 小大與獎助學金
                if sb:
                    row_index += 2
                    sheet.cell(column=1, row=row_index).value = '[小大]'
                    sheet.cell(column=2, row=row_index).value = sb

                if scholarship:
                    row_index += 2
                    sheet.cell(column=1, row=row_index).value = '[子女獎助學金]'
                    sheet.cell(column=2, row=row_index).value = scholarship[:-1]

            row_index += 1
            separate_end = ' ' + '='*129 + ' '
            sheet.cell(column=col_index, row=row_index).value = separate_end
            row_index += 1
            sheet.cell(column=col_index, row=row_index).value = ''

        excel_name = FOLDER_PATH + '\\' + inv_name + '＿主選公務資料.xlsx' if MAIN else FOLDER_PATH + '\\' + inv_name + '＿備選3套公務資料' + '.xlsx'
        wb.save(excel_name)
        output_sample_roster(inv_name, samples)


# 輸出樣本名冊 excel
def output_sample_roster(name, s, type_flag=TYPE_FLAG) -> None:
    flag = False
    inv_name = name
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = inv_name
    row_index = 4
    col_index = 1
    for sample in s:
        if row_index == 4:
            width = list(
                map(lambda x: x * 1.13, [5.29, 5.29, 13.29, 9.29, 9.29, 10.29, 50.29, 4.29, 10.29, 20.29, 5.29]))
            for i in range(1, len(width) + 1):
                sheet.column_dimensions[get_column_letter(i)].width = width[i - 1]

            titles = ['107年主力農家所得調查樣本名冊─' + type_flag, '本頁已完成調查戶數：_____', '失敗戶請填寫失敗訪視紀錄表', '']
            for index, title in enumerate(titles, start=1):
                sheet.merge_cells(start_row=index, start_column=col_index, end_row=index, end_column=11)
                cell = sheet.cell(index, col_index)
                cell.value = title
                cell.alignment = ALIGNMENT
                if index == 3:
                    cell.alignment = Alignment(horizontal='right')
                if index == 4:
                    for i in range(1, 12):
                        sheet.cell(index, i).border = BORDER
        if sample.num in EXCEPT_NUM:
            num = '*' + sample.num
            flag = True
        else:
            num = sample.num
        main_type = sample.main_type
        if main_type.find('(') != -1:
            main_type = main_type[:main_type.index('(')]
        sorted_sample = ['', sample.sample_num, num, sample.num[-5:],
                         sample.name, sample.tel, sample.addr, sample.layer, main_type, sample.area, '']
        if row_index == 4:
            row_index += 1
            set_excel_title(sheet, row_index, 'sample_roster', SAMPLE_ROSTER_TITLES)
        row_index += 1
        sheet.row_dimensions[row_index].height = 1.95 * 16.153
        for index, i in enumerate(sorted_sample, start=1):
            cell = sheet.cell(row_index, index)
            if index in [2, 4, 8]:
                cell.alignment = ALIGNMENT
            if index == 1:
                cell.value = row_index - 5
            else:
                cell.alignment = Alignment(wrap_text=True)
                cell.value = i

            cell.border = BORDER

    if flag:
        row_index += 1
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=11)
        sheet.cell(column=1, row=row_index).value = "備註: 星號(*)為家庭收支的調查對象,請換戶"
    excel_name = FOLDER_PATH + '\\' + inv_name + '＿主選樣本名冊.xlsx' if MAIN else FOLDER_PATH + '\\' + inv_name + '＿備選3套樣本名冊' + '.xlsx'
    wb.save(excel_name)


start_time = time.time()
read_sample()
output_excel()
m, s = divmod(time.time() - start_time, 60)
print(int(m), 'min', round(s, 1), 'sec')